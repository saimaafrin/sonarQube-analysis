import regex as re
import glob
import os
from openpyxl import load_workbook
def task_func(directory_path='./xlsx_files/'):
    # Initialize a counter for the number of files processed
    file_count = 0
    
    # Use glob to find all .xlsx files in the specified directory
    xlsx_files = glob.glob(os.path.join(directory_path, '*.xlsx'))
    
    # Iterate over each file found
    for file_path in xlsx_files:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Iterate over each sheet in the workbook
        for sheet in workbook.sheetnames:
            # Select the sheet
            worksheet = workbook[sheet]
            
            # Iterate over each cell in the sheet
            for row in worksheet.iter_rows():
                for cell in row:
                    # Check if the cell value is a string and contains double quotes
                    if isinstance(cell.value, str) and '"' in cell.value:
                        # Use regex to replace all double quotes with double backslashes followed by double quotes
                        cell.value = re.sub(r'"', r'\"', cell.value)
        
        # Save the workbook with the changes
        workbook.save(file_path)
        
        # Increment the file count
        file_count += 1
    
    # Return the number of files processed
    return file_count