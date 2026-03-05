import os
from pathlib import Path
import pandas as pd
import docx
from openpyxl import load_workbook
def task_func(source_directory: str, target_directory: str) -> int:
    # Check if the source directory exists
    if not os.path.exists(source_directory):
        raise FileNotFoundError(f"Source directory '{source_directory}' does not exist.")
    
    # Initialize a counter for the number of successfully converted files
    converted_count = 0
    
    # Iterate over all files in the source directory
    for file_path in Path(source_directory).glob('*'):
        # Check if the file has one of the specified extensions
        if file_path.suffix in ['.txt', '.docx', '.xlsx', '.csv']:
            # Create the target file path
            target_file_path = Path(target_directory) / file_path.name.replace(file_path.suffix, '.csv')
            
            # Convert the file to CSV
            if file_path.suffix == '.txt':
                # Read the text content from the file
                with open(file_path, 'r', encoding='utf-8') as file:
                    text_content = file.read()
                
                # Create a DataFrame with the text content
                df = pd.DataFrame({'Text': [text_content]})
                
                # Save the DataFrame to a CSV file
                df.to_csv(target_file_path, index=False)
            elif file_path.suffix == '.docx':
                # Read the text content from the docx file
                doc = docx.Document(file_path)
                text_content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
                
                # Create a DataFrame with the text content
                df = pd.DataFrame({'Text': [text_content]})
                
                # Save the DataFrame to a CSV file
                df.to_csv(target_file_path, index=False)
            elif file_path.suffix == '.xlsx':
                # Load the workbook
                workbook = load_workbook(file_path)
                
                # Read the text content from the first sheet
                sheet = workbook.active
                text_content = '\n'.join([cell.value for row in sheet.iter_rows(values_only=True) for cell in row])
                
                # Create a DataFrame with the text content
                df = pd.DataFrame({'Text': [text_content]})
                
                # Save the DataFrame to a CSV file
                df.to_csv(target_file_path, index=False)
            elif file_path.suffix == '.csv':
                # Read the CSV file
                df = pd.read_csv(file_path)
                
                # Save the DataFrame to a CSV file (no change)
                df.to_csv(target_file_path, index=False)
            
            # Increment the counter
            converted_count += 1
    
    # Return the number of successfully converted files
    return converted_count