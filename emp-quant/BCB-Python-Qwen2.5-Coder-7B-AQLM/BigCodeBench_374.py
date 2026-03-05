
import regex as re
import glob
import os
from openpyxl import load_workbook

def task_func(directory_path='./xlsx_files/'):
    processed_files = 0
    # Iterate over all .xlsx files in the specified directory
    for file_path in glob.glob(os.path.join(directory_path, '*.xlsx')):
        try:
            # Load the workbook
            workbook = load_workbook(file_path)
            # Iterate over all sheets in the workbook
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Iterate over all cells in the sheet
                for row in sheet.iter_rows(values_only=False):
                    for cell in row:
                        if isinstance(cell.value, str):
                            # Replace double quotes with double backslashes
                            cell.value = re.sub(r'"', r'\"', cell.value)
            # Save the workbook with the changes
            workbook.save(file_path)
            processed_files += 1
        except Exception as e:
            print(f"Error processing {file_path}: {e}")
    return processed_files