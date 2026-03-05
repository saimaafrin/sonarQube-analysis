
import os
from pathlib import Path
import pandas as pd
import docx
from openpyxl import load_workbook

def task_func(source_directory: str, target_directory: str) -> int:
    # Check if the source directory exists
    source_path = Path(source_directory)
    if not source_path.exists():
        raise FileNotFoundError(f"The source directory '{source_directory}' does not exist.")
    
    # Create the target directory if it does not exist
    target_path = Path(target_directory)
    target_path.mkdir(parents=True, exist_ok=True)
    
    # Initialize a counter for the number of successfully converted files
    converted_count = 0
    
    # Iterate over all files in the source directory
    for file_path in source_path.iterdir():
        if file_path.is_file():
            # Determine the file extension
            file_extension = file_path.suffix.lower()
            
            # Convert the file to CSV if it has one of the specified extensions
            if file_extension in ['.txt', '.docx', '.xlsx', '.csv']:
                # Create the target CSV file path
                csv_file_path = target_path / (file_path.stem + '.csv')
                
                # Read the content of the file based on its extension
                if file_extension == '.txt':
                    with open(file_path, 'r', encoding='utf-8') as file:
                        content = file.read()
                elif file_extension == '.docx':
                    doc = docx.Document(file_path)
                    content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
                elif file_extension == '.xlsx':
                    workbook = load_workbook(file_path)
                    sheet = workbook.active
                    content = '\n'.join([cell.value for row in sheet.iter_rows(values_only=True) for cell in row])
                elif file_extension == '.csv':
                    content = pd.read_csv(file_path, header=None).to_csv(index=False)
                
                # Write the content to the CSV file
                with open(csv_file_path, 'w', encoding='utf-8') as csv_file:
                    csv_file.write(content)
                
                # Increment the counter for successfully converted files
                converted_count += 1
    
    # Return the number of successfully converted files
    return converted_count