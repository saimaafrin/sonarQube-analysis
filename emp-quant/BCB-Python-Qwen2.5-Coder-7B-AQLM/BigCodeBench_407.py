
import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path) -> str:
    # Load the workbook
    workbook = load_workbook(excel_file_path)
    # Select the active sheet
    sheet = workbook.active
    
    # Create the CSV file path
    csv_file_path = os.path.join(csv_file_path, f"{file_name}.csv")
    
    # Open the CSV file for writing
    with open(csv_file_path, mode='w', newline='') as csv_file:
        csv_writer = csv.writer(csv_file)
        
        # Iterate through each row in the sheet
        for row in sheet.iter_rows(values_only=True):
            csv_writer.writerow(row)
    
    # Return the name of the created CSV file
    return os.path.basename(csv_file_path)