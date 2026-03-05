
import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path):
    # Check if the input file exists
    if not os.path.isfile(excel_file_path):
        raise ValueError("File not found")

    # Load the Excel file
    wb = load_workbook(excel_file_path)

    # Get the active sheet
    ws = wb.active

    # Define the CSV file name
    csv_file_name = file_name + ".csv"

    # Write the CSV file
    with open(csv_file_path + csv_file_name, "w", newline="") as csv_file:
        csv_writer = csv.writer(csv_file)
        for row in ws.iter_rows(values_only=True):
            csv_writer.writerow(row)

    return csv_file_name