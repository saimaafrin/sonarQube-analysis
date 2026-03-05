
import regex as re
import glob
import os
from openpyxl import load_workbook
def task_func(directory_path='./xlsx_files/'):
    # Get all Excel files in the specified directory
    excel_files = glob.glob(directory_path + '*.xlsx')
    # Initialize the number of Excel files processed
    num_processed = 0
    # Loop through each Excel file
    for excel_file in excel_files:
        # Load the Excel file using openpyxl
        wb = load_workbook(excel_file)
        # Get the number of worksheets in the Excel file
        num_worksheets = wb.sheet_names.count()
        # Loop through each worksheet in the Excel file
        for worksheet_name in wb.sheet_names:
            # Get the worksheet object
            worksheet = wb[worksheet_name]
            # Get the number of rows in the worksheet
            num_rows = worksheet.max_row
            # Loop through each row in the worksheet
            for row in range(1, num_rows + 1):
                # Get the cell object at the specified row and column
                cell = worksheet[row, 1]
                # Check if the cell contains a double quote
                if re.search(r'"', cell.value):
                    # Prefix the double quote with a double backslash
                    cell.value = '\\"' + cell.value
                # Save the Excel file
                wb.save(excel_file)
                # Increment the number of Excel files processed
                num_processed += 1
    # Return the number of Excel files processed
    return num_processed