import regex as re
import glob
import os
from openpyxl import load_workbook
def task_func(directory_path='./xlsx_files/'):
    # Initialize a counter for processed files
    processed_files = 0
    
    # Use glob to find all .xlsx files in the specified directory
    for file_path in glob.glob(os.path.join(directory_path, '*.xlsx')):
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Iterate through all sheets in the workbook
        for sheet in workbook.sheetnames:
            # Select the sheet
            worksheet = workbook[sheet]
            
            # Iterate through all rows in the sheet
            for row in worksheet.iter_rows(values_only=True):
                # Iterate through all cells in the row
                for cell in row:
                    # Check if the cell value is a string and contains double quotes
                    if isinstance(cell, str) and '"' in cell:
                        # Replace double quotes with double backslashes
                        cell = re.sub(r'"', r'\"', cell)
                        # Update the cell value
                        worksheet[cell.coordinate] = cell
        
        # Save the workbook
        workbook.save(file_path)
        # Increment the counter
        processed_files += 1
    
    # Return the number of processed files
    return processed_files