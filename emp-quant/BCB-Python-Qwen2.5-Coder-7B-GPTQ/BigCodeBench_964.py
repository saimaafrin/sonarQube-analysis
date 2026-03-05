import os
from pathlib import Path
import pandas as pd
import docx
from openpyxl import load_workbook
def task_func(source_directory: str, target_directory: str) -> int:
    # Check if the source directory exists
    if not os.path.exists(source_directory):
        raise FileNotFoundError("Source directory does not exist")
    
    # Initialize a counter for successful conversions
    successful_conversions = 0
    
    # Iterate over each file in the source directory
    for file in os.listdir(source_directory):
        # Construct the full file path
        file_path = os.path.join(source_directory, file)
        
        # Check if the file is a text file
        if file.endswith('.txt'):
            # Read the text content
            with open(file_path, 'r', encoding='utf-8') as f:
                text_content = f.read()
            
            # Create a DataFrame with the text content
            df = pd.DataFrame({'Text': [text_content]})
            
            # Construct the target file path
            target_file_path = os.path.join(target_directory, file.replace('.txt', '.csv'))
            
            # Save the DataFrame to a CSV file
            df.to_csv(target_file_path, index=False)
            successful_conversions += 1
        
        # Check if the file is a Word document
        elif file.endswith('.docx'):
            # Load the Word document
            doc = docx.Document(file_path)
            
            # Extract text from the document
            text_content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
            
            # Create a DataFrame with the text content
            df = pd.DataFrame({'Text': [text_content]})
            
            # Construct the target file path
            target_file_path = os.path.join(target_directory, file.replace('.docx', '.csv'))
            
            # Save the DataFrame to a CSV file
            df.to_csv(target_file_path, index=False)
            successful_conversions += 1
        
        # Check if the file is an Excel file
        elif file.endswith('.xlsx'):
            # Load the Excel file
            workbook = load_workbook(file_path)
            
            # Extract text from the first sheet
            sheet = workbook.active
            text_content = '\n'.join([cell.value for row in sheet.iter_rows(values_only=True) for cell in row])
            
            # Create a DataFrame with the text content
            df = pd.DataFrame({'Text': [text_content]})
            
            # Construct the target file path
            target_file_path = os.path.join(target_directory, file.replace('.xlsx', '.csv'))
            
            # Save the DataFrame to a CSV file
            df.to_csv(target_file_path, index=False)
            successful_conversions += 1
        
        # Check if the file is a CSV file
        elif file.endswith('.csv'):
            # Skip conversion for existing CSV files
            continue
    
    return successful_conversions