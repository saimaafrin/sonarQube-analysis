import os
import csv
from openpyxl import load_workbook
def task_func(file_name, excel_file_path, csv_file_path) -> str:
    # Load the workbook from the specified Excel file path
    workbook = load_workbook(excel_file_path)
    # Select the active sheet
    sheet = workbook.active
    
    # Construct the full path for the CSV file
    csv_file_path = os.path.join(csv_file_path, f"{file_name}.csv")
    
    # Open a CSV file for writing
    with open(csv_file_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        
        # Iterate through each row in the active sheet
        for row in sheet.iter_rows(values_only=True):
            # Write the row to the CSV file
            writer.writerow(row)
    
    # Return the path to the created CSV file
    return csv_file_path