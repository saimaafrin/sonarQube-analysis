import os
from pathlib import Path
import pandas as pd
import docx
from openpyxl import load_workbook
def task_func(source_directory: str, target_directory: str) -> int:
    # Check if the source directory exists
    source_path = Path(source_directory)
    if not source_path.exists():
        raise FileNotFoundError(f"The source directory {source_directory} does not exist.")
    
    # Create the target directory if it doesn't exist
    target_path = Path(target_directory)
    target_path.mkdir(parents=True, exist_ok=True)
    
    # List of supported file extensions
    supported_extensions = ['.txt', '.docx', '.xlsx', '.csv']
    
    # Initialize a counter for successful conversions
    successful_conversions = 0
    
    # Iterate over all files in the source directory
    for file_path in source_path.iterdir():
        if file_path.is_file() and file_path.suffix in supported_extensions:
            # Read the content of the file based on its extension
            if file_path.suffix == '.txt':
                with open(file_path, 'r', encoding='utf-8') as file:
                    content = file.read()
            elif file_path.suffix == '.docx':
                doc = docx.Document(file_path)
                content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
            elif file_path.suffix == '.xlsx':
                workbook = load_workbook(file_path)
                sheet = workbook.active
                content = '\n'.join([str(cell.value) for row in sheet.iter_rows(values_only=True) for cell in row])
            elif file_path.suffix == '.csv':
                df = pd.read_csv(file_path)
                content = '\n'.join(df.values.flatten().astype(str))
            
            # Create a CSV file in the target directory with the same name
            csv_file_path = target_path / (file_path.stem + '.csv')
            df = pd.DataFrame({'Text': [content]})
            df.to_csv(csv_file_path, index=False)
            
            # Increment the counter for successful conversions
            successful_conversions += 1
    
    return successful_conversions