import os
import csv
from openpyxl import load_workbook
def task_func(file_name, excel_file_path, csv_file_path) -> str:
    """
    Converts an Excel file (.xls or .xlsx) to a CSV file by reading the contents of the Excel file and writing them to a new CSV file with the same name but a different extension. Allows specifying separate paths for the Excel file source and the CSV file destination.
    """
    # Check if the Excel file exists
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file {excel_file_path} does not exist")

    # Check if the CSV file exists
    if os.path.exists(csv_file_path):
        raise FileExistsError(f"CSV file {csv_file_path} already exists")

    # Load the Excel file
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Create the CSV file
    with open(csv_file_path, "w", newline="") as csv_file:
        csv_writer = csv.writer(csv_file)

        # Write the header row
        csv_writer.writerow(ws.rows[0])

        # Write the data rows
        for row in ws.iter_rows(values_only=True):
            csv_writer.writerow(row)

    # Return the name of the created CSV file
    return csv_file_path
file_name = "example.xlsx"
excel_file_path = "example.xlsx"
csv_file_path = "example.csv"