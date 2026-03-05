import regex as re
import glob
import os
from openpyxl import load_workbook
def task_func(directory_path='./xlsx_files/'):
    """
    Protects all double quotes in all Excel (.xlsx) files in the specified directory by prefixing them with a double backslash.
    """
    # Get all Excel files in the specified directory
    files = glob.glob(os.path.join(directory_path, '*.xlsx'))
    # Initialize the number of processed files
    num_files = 0
    # Iterate over the Excel files
    for file in files:
        # Open the Excel file
        wb = load_workbook(file)
        # Get the number of worksheets in the Excel file
        num_sheets = len(wb.worksheets)
        # Iterate over the worksheets
        for sheet in wb.worksheets:
            # Get the number of rows in the worksheet
            num_rows = sheet.max_row
            # Iterate over the rows
            for row in range(1, num_rows + 1):
                # Get the number of cells in the row
                num_cells = sheet.max_column
                # Iterate over the cells
                for cell in range(1, num_cells + 1):
                    # Get the cell value
                    value = sheet.cell(row, cell).value
                    # Check if the cell value is a string
                    if isinstance(value, str):
                        # Replace double quotes with double backslashes
                        value = re.sub(r'"', '\\"', value)
                        # Set the new cell value
                        sheet.cell(row, cell).value = value
        # Save the modified Excel file
        wb.save(file)
        # Increment the number of processed files
        num_files += 1
    # Return the number of processed files
    return num_files
directory_path = './xlsx_files/'