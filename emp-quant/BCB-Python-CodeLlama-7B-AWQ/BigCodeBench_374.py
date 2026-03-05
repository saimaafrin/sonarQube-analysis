import regex as re
import glob
import os
from openpyxl import load_workbook
def task_func(directory_path='./xlsx_files/'):
    """
    Protects all double quotes in all Excel (.xlsx) files in the specified directory by prefixing them with a double backslash.
    """
    # Get a list of all Excel files in the specified directory
    xlsx_files = glob.glob(os.path.join(directory_path, '*.xlsx'))

    # Initialize a counter for the number of Excel files processed
    num_processed = 0

    # Iterate over the Excel files
    for xlsx_file in xlsx_files:
        # Open the Excel file using openpyxl
        wb = load_workbook(xlsx_file)

        # Iterate over the worksheets in the Excel file
        for sheet in wb:
            # Iterate over the cells in the worksheet
            for row in sheet.iter_rows():
                for cell in row:
                    # Check if the cell value is a string
                    if isinstance(cell.value, str):
                        # Replace double quotes with double backslashes
                        cell.value = re.sub(r'"', '\\"', cell.value)

        # Save the modified Excel file
        wb.save(xlsx_file)

        # Increment the number of Excel files processed
        num_processed += 1

    # Return the number of Excel files processed
    return num_processed
directory_path = './xlsx_files/'