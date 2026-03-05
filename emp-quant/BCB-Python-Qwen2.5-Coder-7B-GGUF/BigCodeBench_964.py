
import os
from pathlib import Path
import pandas as pd
import docx
from openpyxl import load_workbook

def task_func(source_directory: str, target_directory: str) -> int:
    # Check if the source directory exists
    if not os.path.exists(source_directory):
        raise FileNotFoundError(f"The source directory {source_directory} does not exist.")
    
    # Initialize the count of successfully converted files
    converted_count = 0
    
    # Iterate over each file in the source directory
    for file_name in os.listdir(source_directory):
        # Construct the full file path
        source_file_path = os.path.join(source_directory, file_name)
        
        # Check if the file is a text file
        if file_name.endswith('.txt'):
            # Read the text content of the file
            with open(source_file_path, 'r', encoding='utf-8') as file:
                text_content = file.read()
            
            # Create a DataFrame with the text content
            df = pd.DataFrame({'Text': [text_content]})
            
            # Construct the target file path
            target_file_path = os.path.join(target_directory, file_name.replace('.txt', '.csv'))
            
            # Save the DataFrame to a CSV file
            df.to_csv(target_file_path, index=False)
            converted_count += 1
        
        # Check if the file is a Word document
        elif file_name.endswith('.docx'):
            # Load the Word document
            doc = docx.Document(source_file_path)
            
            # Extract the text content from the document
            text_content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
            
            # Create a DataFrame with the text content
            df = pd.DataFrame({'Text': [text_content]})
            
            # Construct the target file path
            target_file_path = os.path.join(target_directory, file_name.replace('.docx', '.csv'))
            
            # Save the DataFrame to a CSV file
            df.to_csv(target_file_path, index=False)
            converted_count += 1
        
        # Check if the file is an Excel file
        elif file_name.endswith('.xlsx'):
            # Load the Excel file
            workbook = load_workbook(source_file_path)
            
            # Extract the text content from the first sheet
            sheet = workbook.active
            text_content = '\n'.join([cell.value for row in sheet.iter_rows(values_only=True) for cell in row])
            
            # Create a DataFrame with the text content
            df = pd.DataFrame({'Text': [text_content]})
            
            # Construct the target file path
            target_file_path = os.path.join(target_directory, file_name.replace('.xlsx', '.csv'))
            
            # Save the DataFrame to a CSV file
            df.to_csv(target_file_path, index=False)
            converted_count += 1
        
        # Check if the file is a CSV file
        elif file_name.endswith('.csv'):
            # Construct the target file path
            target_file_path = os.path.join(target_directory, file_name)
            
            # Copy the file to the target directory
            os.rename(source_file_path, target_file_path)
            converted_count += 1
    
    return converted_count