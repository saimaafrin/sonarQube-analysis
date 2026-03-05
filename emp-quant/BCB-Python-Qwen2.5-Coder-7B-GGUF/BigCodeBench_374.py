
import regex as re
import glob
import os
from openpyxl import load_workbook

def task_func(directory_path='./xlsx_files/'):
    # Initialize the count of processed files
    processed_files = 0
    
    # Use glob to find all .xlsx files in the specified directory
    xlsx_files = glob.glob(os.path.join(directory_path, '*.xlsx'))
    
    # Loop through each file found
    for file_path in xlsx_files:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Loop through each sheet in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Loop through each cell in the sheet
            for row in sheet.iter_rows():
                for cell in row:
                    # Check if the cell value is a string and contains double quotes
                    if isinstance(cell.value, str) and '"' in cell.value:
                        # Replace double quotes with double backslashes
                        cell.value = re.sub(r'"', r'\"', cell.value)
        
        # Save the workbook with the changes
        workbook.save(file_path)
        
        # Increment the count of processed files
        processed_files += 1
    
    # Return the number of processed files
    return processed_files