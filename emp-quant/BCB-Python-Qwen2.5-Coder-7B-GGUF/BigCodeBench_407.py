
import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path) -> str:
    # Load the workbook from the specified Excel file path
    workbook = load_workbook(excel_file_path)
    
    # Select the active sheet
    sheet = workbook.active
    
    # Create the CSV file path by replacing the file extension
    csv_file_path = os.path.join(csv_file_path, os.path.splitext(file_name)[0] + '.csv')
    
    # Open the CSV file for writing
    with open(csv_file_path, mode='w', newline='') as csv_file:
        csv_writer = csv.writer(csv_file)
        
        # Iterate through each row in the sheet and write to the CSV file
        for row in sheet.iter_rows(values_only=True):
            csv_writer.writerow(row)
    
    # Return the name of the created CSV file
    return os.path.basename(csv_file_path)