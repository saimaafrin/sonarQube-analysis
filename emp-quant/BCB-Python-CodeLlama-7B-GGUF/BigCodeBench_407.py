
import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path) -> str:
    # Check if the input file exists
    if not os.path.isfile(excel_file_path):
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    # Open the Excel file
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Get the sheet name
    sheet_name = ws.title

    # Create the CSV file
    csv_file_name = f"{sheet_name}.csv"
    csv_file_path = os.path.join(csv_file_path, csv_file_name)
    with open(csv_file_path, "w", newline="") as csv_file:
        writer = csv.writer(csv_file)

        # Write the header row
        header_row = [cell.value for cell in ws[ws.dimensions[0]]]
        writer.writerow(header_row)

        # Write the data rows
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    # Return the CSV file name
    return csv_file_name