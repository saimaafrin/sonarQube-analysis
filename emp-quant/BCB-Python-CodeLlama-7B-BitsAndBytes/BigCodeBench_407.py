
import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path):
    # Open the Excel file
    wb = load_workbook(filename=excel_file_path)
    # Get the active sheet
    sheet = wb.active
    # Create a CSV writer
    csv_writer = csv.writer(open(csv_file_path, 'w', newline=''))
    # Write the header row
    csv_writer.writerow(sheet.row_values(0))
    # Write the data rows
    for i in range(1, sheet.max_row + 1):
        csv_writer.writerow(sheet.row_values(i))
    # Close the CSV file
    csv_writer.close()
    # Return the name of the created CSV file
    return os.path.splitext(file_name)[0] + '.csv'