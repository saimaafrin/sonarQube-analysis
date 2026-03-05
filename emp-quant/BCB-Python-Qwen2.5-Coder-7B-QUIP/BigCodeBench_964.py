
import os
from pathlib import Path
import pandas as pd
import docx
from openpyxl import load_workbook

def task_func(source_directory: str, target_directory: str) -> int:
    # Check if the source directory exists
    if not Path(source_directory).exists():
        raise FileNotFoundError(f"Source directory {source_directory} does not exist.")

    # Initialize a counter for successfully converted files
    converted_count = 0

    # Iterate over the files in the source directory
    for file_name in os.listdir(source_directory):
        # Check if the file has one of the specified extensions
        if file_name.endswith(('.txt', '.docx', '.xlsx', '.csv')):
            # Construct the full file path
            source_file_path = Path(source_directory) / file_name

            # Read the content of the file
            if file_name.endswith('.txt'):
                with open(source_file_path, 'r') as file:
                    content = file.read()
            elif file_name.endswith('.docx'):
                doc = docx.Document(source_file_path)
                content = '\n'.join([para.text for para in doc.paragraphs])
            elif file_name.endswith('.xlsx'):
                wb = load_workbook(source_file_path)
                content = '\n'.join([str(cell.value) for cell in wb.active.iter_rows()])[1:]
            elif file_name.endswith('.csv'):
                df = pd.read_csv(source_file_path)
                content = df.to_string(index=False)

            # Construct the target file path
            target_file_path = Path(target_directory) / file_name.replace('.xlsx', '.csv')

            # Write the content to a CSV file
            with open(target_file_path, 'w') as file:
                file.write(content)

            # Increment the counter
            converted_count += 1

    return converted_count