
import regex as re
import glob
import os
from openpyxl import load_workbook

def task_func(directory_path='./xlsx_files/'):
    # Initialize a counter for the number of files processed
    file_count = 0

    # Get a list of all .xlsx files in the specified directory
    xlsx_files = glob.glob(os.path.join(directory_path, '*.xlsx'))

    # Iterate over each .xlsx file
    for file in xlsx_files:
        # Load the workbook
        workbook = load_workbook(file, data_only=True)

        # Iterate over each sheet in the workbook
        for sheet in workbook.sheetnames:
            # Load the sheet
            worksheet = workbook[sheet]

            # Iterate over each cell in the sheet
            for row in worksheet.iter_rows():
                for cell in row:
                    # Check if the cell contains a double quote
                    if '"' in str(cell.value):
                        # Replace double quotes with double backslashes
                        cell.value = re.sub(r'"', r'\"', str(cell.value))

        # Save the workbook with the changes
        workbook.save(file)

        # Increment the file count
        file_count += 1

    # Return the number of files processed
    return file_count