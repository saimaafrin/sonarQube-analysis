
import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path) -> str:
    # Load the workbook
    workbook = load_workbook(excel_file_path, data_only=True)
    # Get the active sheet
    sheet = workbook.active
    
    # Create a CSV file with the same name but different extension
    csv_file_name = os.path.splitext(file_name)[0] + '.csv'
    csv_file_path = os.path.join(csv_file_path, csv_file_name)
    
    # Write the data to the CSV file
    with open(csv_file_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)
    
    # Return the name of the created CSV file
    return csv_file_name