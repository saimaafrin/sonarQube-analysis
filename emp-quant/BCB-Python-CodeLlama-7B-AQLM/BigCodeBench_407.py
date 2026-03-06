import os
import csv
from openpyxl import load_workbook
def task_func(file_name, excel_file_path, csv_file_path) -> str:
    """
    Converts an Excel file (.xls or .xlsx) to a CSV file by reading the contents of the Excel file
    and writing them to a new CSV file with the same name but a different extension. Allows specifying
    separate paths for the Excel file source and the CSV file destination.

    Parameters:
        file_name (str): The name of the Excel file to be converted.
        excel_file_path (str): The directory path where the Excel file is located.
        csv_file_path (str): The directory path where the CSV file should be saved.

    Returns:
        str: The name of the created CSV file.

    Requirements:
    - openpyxl.load_workbook
    - os
    - csv

    Example:
    >>> task_func('test.xlsx', '/path/to/excel/files', '/path/to/csv/files')
    'test.csv'
    >>> task_func('nonexistent.xlsx', '/path/to/excel/files', '/path/to/csv/files')
    Traceback (most recent call last):
       ...
    FileNotFoundError: [Errno 2] No such file or directory: '/path/to/excel/files/nonexistent.xlsx'

    Note:
    - This function assumes the active sheet is the one to be converted.
    """
    try:
        # Open the Excel file
        wb = load_workbook(os.path.join(excel_file_path, file_name))
        # Get the active sheet
        ws = wb.active
        # Get the name of the active sheet
        sheet_name = ws.title
        # Get the number of rows in the active sheet
        num_rows = ws.max_row
        # Get the number of columns in the active sheet
        num_cols = ws.max_column
        # Create a new CSV file with the same name as the Excel file
        csv_file_name = file_name.replace('.xlsx', '.csv')
        csv_file_path = os.path.join(csv_file_path, csv_file_name)
        # Open the CSV file
        csv_file = open(csv_file_path, 'w')
        # Write the header row
        csv_writer = csv.writer(csv_file)
        csv_writer.writerow(ws.rows[0])
        # Write the data rows
        for row in ws.rows[1:]:
            csv_writer.writerow(row)
        # Close the CSV file
        csv_file.close()
        # Return the name of the created CSV file
        return csv_file_name
    except FileNotFoundError:
        # If the Excel file does not exist, raise a FileNotFoundError
        raise FileNotFoundError(f'[Errno 2] No such file or directory: {os.path.join(excel_file_path, file_name)}')
    except PermissionError:
        # If the Excel file is not readable, raise a PermissionError
        raise PermissionError(f'[Errno 13] Permission denied: {os.path.join(excel_file_path, file_name)}')
    except Exception as e:
        # If an unexpected error occurs, raise the error
        raise e