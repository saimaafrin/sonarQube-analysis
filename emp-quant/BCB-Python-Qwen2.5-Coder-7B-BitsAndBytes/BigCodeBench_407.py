
import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path) -> str:
    # Load the workbook from the specified path
    wb = load_workbook(os.path.join(excel_file_path, file_name))
    
    # Select the active sheet
    sheet = wb.active
    
    # Create the full path for the CSV file
    csv_file_path = os.path.join(csv_file_path, os.path.splitext(file_name)[0] + '.csv')
    
    # Open the CSV file for writing
    with open(csv_file_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        
        # Iterate through each row in the sheet
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)
    
    return csv_file_path