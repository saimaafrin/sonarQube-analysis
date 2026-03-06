
import re
import glob
import os
from openpyxl import load_workbook

def task_func(directory_path='./xlsx_files/'):
    # Initialize the count of processed files
    file_count = 0
    
    # Get a list of all .xlsx files in the specified directory
    xlsx_files = glob.glob(os.path.join(directory_path, '*.xlsx'))
    
    for file_path in xlsx_files:
        try:
            # Load the workbook
            wb = load_workbook(file_path)
            
            # Iterate through all sheets in the workbook
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Iterate through all cells in the sheet
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            # Replace double quotes with two backslashes
                            cell.value = re.sub(r'"', r'\"', cell.value)
            
            # Save the modified workbook
            wb.save(file_path)
            
            # Increment the count of processed files
            file_count += 1
        
        except Exception as e:
            print(f"Error processing {file_path}: {e}")
    
    return file_count