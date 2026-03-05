import os
import csv
from openpyxl import load_workbook
def task_func(file_name, excel_file_path, csv_file_path) -> str:
    """
    Converts an Excel file (.xls or .xlsx) to a CSV file by reading the contents of the Excel file and writing them to a new CSV file with the same name but a different extension. Allows specifying separate paths for the Excel file source and the CSV file destination.
    """
    # Check if the input file exists
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"{excel_file_path} does not exist")

    # Check if the output file already exists
    if os.path.exists(csv_file_path):
        raise FileExistsError(f"{csv_file_path} already exists")

    # Load the Excel file
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Create the CSV file
    with open(csv_file_path, "w", newline="") as csv_file:
        writer = csv.writer(csv_file)

        # Write the header row
        header_row = [cell.value for cell in ws[ws.dimensions[0]]]
        writer.writerow(header_row)

        # Write the data rows
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    # Return the name of the created CSV file
    return os.path.basename(csv_file_path)
file_name = "example.xlsx"
excel_file_path = "example.xlsx"
csv_file_path = "example.csv"