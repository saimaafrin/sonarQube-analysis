import regex as re
import glob
import os
from openpyxl import load_workbook
def task_func(directory_path='./xlsx_files/'):
    """
    Protects all double quotes in all Excel (.xlsx) files in the specified directory by prefixing them with a double backslash.
    """
    # Get all Excel files in the specified directory
    excel_files = glob.glob(os.path.join(directory_path, '*.xlsx'))

    # Initialize counter for number of Excel files processed
    num_files_processed = 0

    # Iterate over Excel files
    for excel_file in excel_files:
        # Load Excel file using openpyxl
        wb = load_workbook(excel_file)

        # Iterate over worksheets in Excel file
        for ws in wb:
            # Iterate over cells in worksheet
            for row in ws.iter_rows(values_only=True):
                # Check if cell value is a string
                if isinstance(row[0], str):
                    # Replace double quotes with double backslashes
                    row[0] = re.sub(r'"', '\\"', row[0])

        # Save changes to Excel file
        wb.save(excel_file)

        # Increment counter for number of Excel files processed
        num_files_processed += 1

    # Return number of Excel files processed
    return num_files_processed
directory_path = './xlsx_files/'